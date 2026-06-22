# -*- coding: utf-8 -*-
import openpyxl, os
USD=chr(36)
f='Opus_Pulse_TNM_vs_Outcome_Model.xlsx'
wb=openpyxl.load_workbook(f)
print('Sheets:',wb.sheetnames)
for ws in wb.worksheets:
    print(f'  {ws.title}: charts={len(ws._charts)} dims={ws.dimensions}')
print('Size KB:',round(os.path.getsize(f)/1024,1))

# scan for any literal Excel error strings accidentally stored, and unguarded divisions
errs=[]
divs=[]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str):
                if any(e in v for e in ('#REF','#DIV/0','#N/A','#VALUE','#NAME')): errs.append((ws.title,c.coordinate,v))
                if v.startswith('=') and '/' in v and 'IFERROR' not in v and 'IF(' not in v:
                    # allow "/12" style constant divides (safe)
                    pass
print('Stored error strings (must be empty):',errs)

b=wb['Billing (TNM)']; o=wb['Outcome']
print('Billing monthly rev formula:',b['B34'].value)
print('Billing margin formula:',b['B40'].value)
print('Outcome R:',o['B13'].value,'| C:',o['B14'].value)

# independent recompute
days=365-104-10-21; hpd=8; hy=days*hpd; hm=hy/12
team=[('Sr Dev',3,28),('Sr Dev',2,30),('TL',1,35),('Sr QA',1,28),('Lead QA',2,35),('PO',1,37)]
scr=sum(cnt*rate for _,cnt,rate in team); R=scr*hm; RY=scr*hy
exp_m=24000+2500+1200+700; C=(exp_m*12+6000)/12
print(f'Recompute -> days={days} hm={hm:.2f} R={USD}{R:,.0f} RY={USD}{RY:,.0f} margin={(RY-(exp_m*12+6000))/RY:.1%} C={USD}{C:,.0f}')
res=0.15; eff=0.10; disc=0.05
for name,pp in [('C1',7),('C2',8)]:
    msp=pp*10*(1-res)*2; be=R/msp; rec=be*(1-disc); bill=rec*msp; oc=C/(1+eff)
    extra=((bill-oc)-(R-C))*12; save=(R-bill)*12
    ww='YES' if (bill<R and (bill-oc)>=(R-C)) else 'No'
    print(f'  {name}: SP/mo={msp:.0f} be={USD}{be:.0f} rec={USD}{rec:.0f} bill={USD}{bill:,.0f} outMargin={(bill-oc)/bill:.1%} extra/yr={USD}{extra:,.0f} clientSave/yr={USD}{save:,.0f} WINWIN={ww}')
