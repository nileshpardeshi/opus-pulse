# -*- coding: utf-8 -*-
"""
Opus Pulse - TNM -> Outcome conversion: a polished, formula-driven Excel simulator.
YELLOW cells are inputs; everything else is a live formula. All divisions/lookups
are IFERROR-guarded (no #DIV/0! or #N/A). Charts are styled and placed clear of data.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ── theme ─────────────────────────────────────────────────────────────────────
NAVY="14315A"; BLUE="2F6FB0"; LBLUE="EAF2FB"; SKY="DCE9F7"; GREEN="2E7D32"; GREENL="E7F1E8"
AMBER="C8860D"; AMBERL="FBF1DD"; GREY="6B7480"; LINE="D6DEE8"; YELLOW="FFF4C2"; WHITE="FFFFFF"
BG="F7F9FC"; INK="222B38"; REDL="FAE3E3"
F="Calibri"
MONEY='"$"#,##0'; MONEY2='"$"#,##0.00'; PCT='0.0%'; NUM='#,##0'; NUM1='#,##0.0'
thin=Side(style="thin",color=LINE)
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def fill(c): return PatternFill("solid",fgColor=c)
def font(sz=11,color=INK,bold=False,italic=False): return Font(name=F,size=sz,color=color,bold=bold,italic=italic)
def setc(ws,ref,val=None,*,f=None,bg=None,fmt=None,ha=None,va="center",border=False,wrap=False):
    c=ws[ref]
    if val is not None: c.value=val
    if f: c.font=f
    if bg: c.fill=fill(bg)
    if fmt: c.number_format=fmt
    c.alignment=Alignment(horizontal=ha,vertical=va,wrap_text=wrap)
    if border: c.border=BORDER
    return c
def banner(ws,last,title,sub):
    ws.merge_cells(f"A1:{last}1"); ws.merge_cells(f"A2:{last}2")
    setc(ws,"A1",title,f=font(16,WHITE,True),bg=NAVY,ha="left"); ws.row_dimensions[1].height=26
    setc(ws,"A2",sub,f=font(10,WHITE,italic=True),bg=NAVY,ha="left"); ws.row_dimensions[2].height=16
    # extend banner fill across merged row
    for col in range(1,_colnum(last)+1):
        ws.cell(row=1,column=col).fill=fill(NAVY); ws.cell(row=2,column=col).fill=fill(NAVY)
def _colnum(letter):
    n=0
    for ch in letter: n=n*26+(ord(ch)-64)
    return n
def section(ws,row,last,text):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=_colnum(last))
    setc(ws,f"A{row}",text,f=font(11,WHITE,True),bg=BLUE,ha="left"); ws.row_dimensions[row].height=18
    for col in range(1,_colnum(last)+1): ws.cell(row=row,column=col).fill=fill(BLUE)
def header(ws,row,headers,start=1,fillc=NAVY):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h); c.font=font(11,WHITE,True); c.fill=fill(fillc)
        c.alignment=Alignment(horizontal="left" if i==0 else "center",vertical="center",wrap_text=True); c.border=BORDER
def inp(ws,ref,val,fmt=None,ha="center"):
    return setc(ws,ref,val,f=font(11,INK,True),bg=YELLOW,fmt=fmt,ha=ha,border=True)

wb=Workbook()

# ═══════════════════════════════ 1) WELCOME ═══════════════════════════════════
ws=wb.active; ws.title="Welcome"; ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=NAVY
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=98
banner(ws,"C","Opus Pulse - TNM to Outcome Conversion Simulator","A live model. Change any YELLOW input and every result, table and chart recalculates instantly.")
rows=[("S","How to use"),
      ("Yellow cells","are the only inputs you edit - team, rates, calendar, expenses, capacity, efficiency and discount."),
      ("Everything else","is a live, IFERROR-guarded formula - no #DIV/0! or #N/A, even mid-edit."),
      ("Charts","read from the formula cells, so they redraw automatically."),
      ("S","The tabs"),
      ("Rate Card","Standard billing rate per role and experience. Billing looks rates up from here."),
      ("Billing (TNM)","Time & Material: team + calendar + expenses -> monthly & yearly billing, profit and margin."),
      ("Outcome","Outcome conversion: two capacity cases, price per story point, and the client win-win."),
      ("Dashboard","One-screen verdict: TNM vs Outcome and who benefits."),
      ("S","Key formulas"),
      ("Billable days/yr","365 - 104 weekends - holidays - leaves"),
      ("Monthly billing","sum of (count x rate) x billable hours per month"),
      ("Break-even price/SP","TNM monthly billing / deliverable SP per month"),
      ("Recommended price/SP","break-even x (1 - client discount %)"),
      ("Outcome cost/month","TNM monthly cost / (1 + efficiency gain %)"),
      ("Win-win test","client pays less than TNM AND Opus profit is not lower")]
r=4
for tag,txt in rows:
    if tag=="S": section(ws,r,"C",txt)
    else:
        setc(ws,f"B{r}",tag,f=font(11,NAVY,True),bg=BG,ha="left",border=True)
        setc(ws,f"C{r}",txt,f=font(11,INK),ha="left",border=True,wrap=True)
        ws.row_dimensions[r].height=22 if len(txt)<95 else 30
    r+=1
setc(ws,"B"+str(r+1),"Legend:",f=font(10,GREY,True))
setc(ws,"C"+str(r+1),"yellow = input    |    white/blue = formula (do not edit)",f=font(10,GREY,italic=True),bg=YELLOW,ha="left")

# ═══════════════════════════════ 2) RATE CARD ═════════════════════════════════
ws=wb.create_sheet("Rate Card"); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=BLUE
for col,w in [("A",22),("B",16),("C",16),("D",2),("E",2),("F",22)]: ws.column_dimensions[col].width=w
banner(ws,"C","Account Rate Card  -  Apex Bank","Standard billing rate by role and experience. YELLOW = editable.")
header(ws,4,["Role","Experience (yrs)","Billing Rate/hr"])
rate_rows=[("Sr Dev","5 to 8",28),("Sr Dev","8 to 10",30),("Tech Lead (Dev)","9 to 12",35),
           ("Sr QA","5 to 8",28),("Lead QA","9 to 12",35),("PO","12 to 14",37),
           ("Developer","3 to 5",22),("QA Engineer","3 to 5",20),("Architect","12 to 14",45),("Business Analyst","5 to 8",30)]
RC0=5
for i,(role,exp,rate) in enumerate(rate_rows):
    rr=RC0+i; bf=BG if i%2 else WHITE
    setc(ws,f"A{rr}",role,f=font(11,INK),bg=bf,ha="left",border=True)
    setc(ws,f"B{rr}",exp,f=font(11,INK),bg=bf,ha="center",border=True)
    inp(ws,f"C{rr}",rate,MONEY2)
    setc(ws,f"F{rr}",f'=A{rr}&"|"&B{rr}',f=font(9,GREY))
RCn=RC0+len(rate_rows)-1
ws.column_dimensions["F"].hidden=True
RATE=f"'Rate Card'!$C${RC0}:$C${RCn}"; KEY=f"'Rate Card'!$F${RC0}:$F${RCn}"
ws.freeze_panes="A5"

# ═══════════════════════════════ 3) BILLING (TNM) ═════════════════════════════
ws=wb.create_sheet("Billing (TNM)"); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor="375623"
for col,w in [("A",30),("B",15),("C",11),("D",13),("E",16),("F",16),("G",2),("H",17),("I",13)]: ws.column_dimensions[col].width=w
banner(ws,"F","Billing Simulator - Time & Material (TNM)","YELLOW = inputs. Revenue, profit and margin are live formulas.")
section(ws,4,"C","Company calendar & hours")
calc=[("Hours per day",8,True,NUM),("Weekend days / year",104,False,NUM),
      ("Public holidays / year",10,True,NUM),("Leaves / year",21,True,NUM),
      ("Billable days / year","=365-B6-B7-B8",False,NUM),
      ("Billable hours / year (per resource)","=B9*B5",False,NUM),
      ("Billable hours / month (per resource)","=B10/12",False,NUM1)]
for i,(lbl,val,isin,fmt) in enumerate(calc):
    rr=5+i
    setc(ws,f"A{rr}",lbl,f=font(11,INK),bg=(WHITE if isin else BG),ha="left",border=True)
    if isin: inp(ws,f"B{rr}",val,fmt)
    else: setc(ws,f"B{rr}",val,f=font(11,INK,True),bg=BG,fmt=fmt,ha="center",border=True)
HM="$B$11"; HY="$B$10"
section(ws,13,"F","Project team   (pick role & experience - rate auto-fills from the Rate Card)")
header(ws,14,["Role","Experience","Count","Rate/hr","Monthly Revenue","Yearly Revenue"])
res=[("Sr Dev","5 to 8",3),("Sr Dev","8 to 10",2),("Tech Lead (Dev)","9 to 12",1),
     ("Sr QA","5 to 8",1),("Lead QA","9 to 12",2),("PO","12 to 14",1)]
R0=15
for i,(role,exp,cnt) in enumerate(res):
    rr=R0+i; bf=BG if i%2 else WHITE
    setc(ws,f"A{rr}",role,f=font(11,INK),bg=YELLOW,ha="left",border=True)
    setc(ws,f"B{rr}",exp,f=font(11,INK),bg=YELLOW,ha="center",border=True)
    inp(ws,f"C{rr}",cnt,NUM)
    setc(ws,f"D{rr}",f'=IFERROR(INDEX({RATE},MATCH(A{rr}&"|"&B{rr},{KEY},0)),0)',f=font(11,INK),bg=bf,fmt=MONEY2,ha="center",border=True)
    setc(ws,f"E{rr}",f"=C{rr}*D{rr}*{HM}",f=font(11,INK),bg=bf,fmt=MONEY,ha="center",border=True)
    setc(ws,f"F{rr}",f"=C{rr}*D{rr}*{HY}",f=font(11,INK),bg=bf,fmt=MONEY,ha="center",border=True)
Rn=R0+len(res)-1; TR=Rn+1
ws.merge_cells(f"A{TR}:B{TR}"); setc(ws,f"A{TR}","Total team",f=font(11,NAVY,True),bg=LBLUE,ha="left",border=True)
setc(ws,f"B{TR}",None,bg=LBLUE,border=True)
setc(ws,f"C{TR}",f"=SUM(C{R0}:C{Rn})",f=font(11,NAVY,True),bg=LBLUE,fmt=NUM,ha="center",border=True)
setc(ws,f"D{TR}",None,bg=LBLUE,border=True)
setc(ws,f"E{TR}",f"=SUM(E{R0}:E{Rn})",f=font(11,NAVY,True),bg=LBLUE,fmt=MONEY,ha="center",border=True)
setc(ws,f"F{TR}",f"=SUM(F{R0}:F{Rn})",f=font(11,NAVY,True),bg=LBLUE,fmt=MONEY,ha="center",border=True)
COUNT=f"C{TR}"
E0=TR+2
section(ws,E0,"F","Project expenses (cost)")
header(ws,E0+1,["Category","Amount","Frequency","Yearly Amount"])
exp=[("Salaries",24000,"monthly"),("Infrastructure / seats",2500,"monthly"),("Software licenses",1200,"monthly"),("Travel",700,"monthly"),("Training",6000,"yearly")]
EX0=E0+2
for i,(catg,amt,freq) in enumerate(exp):
    rr=EX0+i; bf=BG if i%2 else WHITE
    setc(ws,f"A{rr}",catg,f=font(11,INK),bg=YELLOW,ha="left",border=True)
    inp(ws,f"B{rr}",amt,MONEY)
    setc(ws,f"C{rr}",freq,f=font(11,INK),bg=YELLOW,ha="center",border=True)
    setc(ws,f"D{rr}",f'=IF(C{rr}="monthly",B{rr}*12,B{rr})',f=font(11,INK),bg=bf,fmt=MONEY,ha="center",border=True)
EXn=EX0+len(exp)-1; ETY=EXn+1; ETM=ETY+1
ws.merge_cells(f"A{ETY}:C{ETY}"); setc(ws,f"A{ETY}","Total yearly expenses",f=font(11,NAVY,True),bg=LBLUE,ha="left",border=True)
for col in "BC": setc(ws,f"{col}{ETY}",None,bg=LBLUE,border=True)
setc(ws,f"D{ETY}",f"=SUM(D{EX0}:D{EXn})",f=font(11,NAVY,True),bg=LBLUE,fmt=MONEY,ha="center",border=True)
ws.merge_cells(f"A{ETM}:C{ETM}"); setc(ws,f"A{ETM}","Total monthly expenses",f=font(11,NAVY,True),bg=LBLUE,ha="left",border=True)
for col in "BC": setc(ws,f"{col}{ETM}",None,bg=LBLUE,border=True)
setc(ws,f"D{ETM}",f"=D{ETY}/12",f=font(11,NAVY,True),bg=LBLUE,fmt=MONEY,ha="center",border=True)
RES=ETM+2
section(ws,RES,"B","Results")
results=[("Monthly billing (revenue)",f"=E{TR}",MONEY,WHITE),("Yearly billing (revenue)",f"=F{TR}",MONEY,BG),
         ("Monthly expenses (cost)",f"=D{ETM}",MONEY,WHITE),("Yearly expenses (cost)",f"=D{ETY}",MONEY,BG),
         ("Monthly profit",None,MONEY,WHITE),("Yearly profit",None,MONEY,BG),("Profit margin %",None,PCT,GREENL)]
rmap={}
for i,(lbl,frm,fmt,bf) in enumerate(results):
    rr=RES+1+i; rmap[lbl]=rr
    setc(ws,f"A{rr}",lbl,f=font(11,NAVY,True),bg=bf,ha="left",border=True)
    setc(ws,f"B{rr}",frm if frm else 0,f=font(11,INK,True),bg=bf,fmt=fmt,ha="center",border=True)
setc(ws,f"B{rmap['Monthly profit']}",f"=B{rmap['Monthly billing (revenue)']}-B{rmap['Monthly expenses (cost)']}",f=font(11,INK,True),bg=WHITE,fmt=MONEY,ha="center",border=True)
setc(ws,f"B{rmap['Yearly profit']}",f"=B{rmap['Yearly billing (revenue)']}-B{rmap['Yearly expenses (cost)']}",f=font(11,INK,True),bg=BG,fmt=MONEY,ha="center",border=True)
setc(ws,f"B{rmap['Profit margin %']}",f"=IFERROR(B{rmap['Yearly profit']}/B{rmap['Yearly billing (revenue)']},0)",f=font(14,GREEN,True),bg=GREENL,fmt=PCT,ha="center",border=True)
BILL_REV=f"'Billing (TNM)'!$B${rmap['Monthly billing (revenue)']}"; BILL_EXP=f"'Billing (TNM)'!$B${rmap['Monthly expenses (cost)']}"
# chart data (clear of tables - column H, above the charts)
setc(ws,"H4","Monthly",f=font(10,NAVY,True)); setc(ws,"I4","$",f=font(10,NAVY,True))
setc(ws,"H5","Revenue"); setc(ws,"I5",f"=B{rmap['Monthly billing (revenue)']}",fmt=MONEY)
setc(ws,"H6","Expenses"); setc(ws,"I6",f"=B{rmap['Monthly expenses (cost)']}",fmt=MONEY)
setc(ws,"H7","Profit"); setc(ws,"I7",f"=B{rmap['Monthly profit']}",fmt=MONEY)
# validations
dv_h=DataValidation(type="list",formula1='"8,9"',allow_blank=False); ws.add_data_validation(dv_h); dv_h.add("B5")
dv_f=DataValidation(type="list",formula1='"monthly,yearly"',allow_blank=False); ws.add_data_validation(dv_f); dv_f.add(f"C{EX0}:C{EXn}")
roles=",".join(sorted(set(r for r,_,_ in rate_rows))); exps=",".join(["0 to 3","3 to 5","5 to 8","8 to 10","9 to 12","12 to 14"])
dv_r=DataValidation(type="list",formula1=f'"{roles}"',allow_blank=True); ws.add_data_validation(dv_r); dv_r.add(f"A{R0}:A{Rn}")
dv_e=DataValidation(type="list",formula1=f'"{exps}"',allow_blank=True); ws.add_data_validation(dv_e); dv_e.add(f"B{R0}:B{Rn}")
# conditional formatting
ws.conditional_formatting.add(f"D{R0}:D{Rn}",CellIsRule(operator="equal",formula=["0"],fill=fill(REDL),font=Font(color="9C2B2B")))
mref=f"B{rmap['Profit margin %']}"
ws.conditional_formatting.add(mref,CellIsRule(operator="greaterThanOrEqual",formula=["0.2"],fill=fill(GREENL),font=Font(color=GREEN,bold=True)))
ws.conditional_formatting.add(mref,CellIsRule(operator="lessThan",formula=["0.15"],fill=fill(REDL),font=Font(color="9C2B2B",bold=True)))
# charts
bar=BarChart(); bar.type="col"; bar.title="Monthly: Revenue / Expenses / Profit"; bar.height=7; bar.width=11; bar.style=10
bar.add_data(Reference(ws,min_col=9,min_row=4,max_row=7),titles_from_data=True); bar.set_categories(Reference(ws,min_col=8,min_row=5,max_row=7))
bar.legend=None; bar.dataLabels=DataLabelList(); bar.dataLabels.showVal=True; bar.dataLabels.numFmt='"$"#,##0'
bar.series[0].graphicalProperties.solidFill=BLUE; bar.y_axis.majorGridlines=None
ws.add_chart(bar,"H9")
pie=PieChart(); pie.title="Expense breakdown (yearly)"; pie.height=7; pie.width=11; pie.style=10
pie.add_data(Reference(ws,min_col=4,min_row=EX0,max_row=EXn)); pie.set_categories(Reference(ws,min_col=1,min_row=EX0,max_row=EXn))
pie.dataLabels=DataLabelList(); pie.dataLabels.showPercent=True
ppal=[BLUE,NAVY,AMBER,"0AA2C0",GREY]
for i in range(len(exp)):
    dp=DataPoint(idx=i); dp.graphicalProperties.solidFill=ppal[i%len(ppal)]; pie.series[0].data_points.append(dp)
ws.add_chart(pie,"H25")
ws.freeze_panes="A3"

# ═══════════════════════════════ 4) OUTCOME ═══════════════════════════════════
ws=wb.create_sheet("Outcome"); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor="7030A0"
for col,w in [("A",32),("B",16),("C",16),("D",2),("N",11),("O",13),("P",13)]: ws.column_dimensions[col].width=w
banner(ws,"C","Outcome-Based Cost Simulation","Two capacity cases. YELLOW = inputs. Team & cost are inherited from Billing (TNM).")
section(ws,4,"C","Inputs")
inp_rows=[("Headcount (from Billing)",f"='Billing (TNM)'!{COUNT}",False,NUM),
          ("Per-person SP / sprint - Case 1",7,True,NUM),("Per-person SP / sprint - Case 2",8,True,NUM),
          ("Capacity reserve %",0.15,True,PCT),("Sprints per month",2,True,NUM),("Minimum invoice SP",120,True,NUM),
          ("Outcome efficiency gain %",0.10,True,PCT),("Client discount %",0.05,True,PCT),
          ("TNM monthly billing  (R)",f"={BILL_REV}",False,MONEY),("Monthly cost  (C)",f"={BILL_EXP}",False,MONEY)]
for i,(lbl,val,isin,fmt) in enumerate(inp_rows):
    rr=5+i
    setc(ws,f"A{rr}",lbl,f=font(11,INK),bg=(WHITE if isin else LBLUE),ha="left",border=True)
    if isin: inp(ws,f"B{rr}",val,fmt)
    else: setc(ws,f"B{rr}",val,f=font(11,NAVY,True),bg=LBLUE,fmt=fmt,ha="center",border=True)
HC="$B$5"; CA="$B$6"; CB="$B$7"; RSV="$B$8"; SPM="$B$9"; MIN="$B$10"; EFF="$B$11"; DIS="$B$12"; RR_="$B$13"; CC="$B$14"
T0=16; section(ws,T0,"C","Case 1 vs Case 2")
header(ws,T0+1,["Metric","Case 1","Case 2"])
def row2(lbl,f1,f2,fmt,hl=False):
    row2.r+=1; rr=row2.r; bf=GREENL if hl else (BG if (rr-T0-2)%2 else WHITE)
    setc(ws,f"A{rr}",lbl,f=font(11,NAVY if hl else INK,hl),bg=bf,ha="left",border=True)
    setc(ws,f"B{rr}",f1,f=font(11,(GREEN if hl else INK),hl),bg=bf,fmt=fmt,ha="center",border=True)
    setc(ws,f"C{rr}",f2,f=font(11,(GREEN if hl else INK),hl),bg=bf,fmt=fmt,ha="center",border=True)
    return rr
row2.r=T0+1
r_pp=row2("Per-person SP / sprint",f"={CA}",f"={CB}",NUM)
r_gv=row2("Gross team velocity / sprint",f"=B{r_pp}*{HC}",f"=C{r_pp}*{HC}",NUM1)
r_ev=row2("Effective velocity (after reserve)",f"=B{r_gv}*(1-{RSV})",f"=C{r_gv}*(1-{RSV})",NUM1)
r_ms=row2("Deliverable SP / month",f"=B{r_ev}*{SPM}",f"=C{r_ev}*{SPM}",NUM)
r_mm=row2("Meets min invoice?",f'=IF(B{r_ms}>={MIN},"Yes","No (short)")',f'=IF(C{r_ms}>={MIN},"Yes","No (short)")',None)
r_be=row2("Break-even price / SP",f"=IFERROR({RR_}/B{r_ms},0)",f"=IFERROR({RR_}/C{r_ms},0)",MONEY2)
r_rc=row2("Recommended price / SP",f"=B{r_be}*(1-{DIS})",f"=C{r_be}*(1-{DIS})",MONEY2,hl=True)
r_mb=row2("Monthly outcome bill",f"=B{r_rc}*B{r_ms}",f"=C{r_rc}*C{r_ms}",MONEY,hl=True)
r_yb=row2("Yearly outcome bill",f"=B{r_mb}*12",f"=C{r_mb}*12",MONEY)
r_oc=row2("Outcome cost / month",f"=IFERROR({CC}/(1+{EFF}),0)",f"=IFERROR({CC}/(1+{EFF}),0)",MONEY)
r_tm=row2("Opus margin - TNM",f"=IFERROR(({RR_}-{CC})/{RR_},0)",f"=IFERROR(({RR_}-{CC})/{RR_},0)",PCT)
r_om=row2("Opus margin - Outcome",f"=IFERROR((B{r_mb}-B{r_oc})/B{r_mb},0)",f"=IFERROR((C{r_mb}-C{r_oc})/C{r_mb},0)",PCT,hl=True)
r_up=row2("Margin uplift",f"=B{r_om}-B{r_tm}",f"=C{r_om}-C{r_tm}",PCT)
r_pt=row2("Opus profit - TNM / month",f"={RR_}-{CC}",f"={RR_}-{CC}",MONEY)
r_po=row2("Opus profit - Outcome / month",f"=B{r_mb}-B{r_oc}",f"=C{r_mb}-C{r_oc}",MONEY)
r_ep=row2("Opus extra profit / year",f"=(B{r_po}-B{r_pt})*12",f"=(C{r_po}-C{r_pt})*12",MONEY,hl=True)
r_cm=row2("Client saving / month",f"={RR_}-B{r_mb}",f"={RR_}-C{r_mb}",MONEY)
r_cy=row2("Client saving / year",f"=B{r_cm}*12",f"=C{r_cm}*12",MONEY,hl=True)
r_ww=row2("WIN-WIN?",f'=IF(AND(B{r_mb}<{RR_},(B{r_mb}-B{r_oc})>=({RR_}-{CC})),"YES - WIN-WIN","Not yet")',
          f'=IF(AND(C{r_mb}<{RR_},(C{r_mb}-C{r_oc})>=({RR_}-{CC})),"YES - WIN-WIN","Not yet")',None,hl=True)
# conditional format win-win green
for cell in (f"B{r_ww}",f"C{r_ww}"):
    ws.conditional_formatting.add(cell,FormulaRule(formula=[f'ISNUMBER(SEARCH("WIN",{cell}))'],fill=fill(GREENL),font=Font(color=GREEN,bold=True)))
    ws.conditional_formatting.add(cell,FormulaRule(formula=[f'NOT(ISNUMBER(SEARCH("WIN",{cell})))'],fill=fill(AMBERL),font=Font(color=AMBER,bold=True)))
# helper data (far-right, hidden) for charts
def hblk(top,title,h1,h2,a1,a2,b1,b2,fmt):
    setc(ws,f"N{top}",title,f=font(10,NAVY,True)); setc(ws,f"O{top}",h1,f=font(10,NAVY,True)); setc(ws,f"P{top}",h2,f=font(10,NAVY,True))
    setc(ws,f"N{top+1}","Case 1"); setc(ws,f"O{top+1}",a1,fmt=fmt); setc(ws,f"P{top+1}",a2,fmt=fmt)
    setc(ws,f"N{top+2}","Case 2"); setc(ws,f"O{top+2}",b1,fmt=fmt); setc(ws,f"P{top+2}",b2,fmt=fmt)
hblk(5,"Client pays","TNM","Outcome",f"={RR_}",f"=B{r_mb}",f"={RR_}",f"=C{r_mb}",MONEY)
hblk(9,"Opus margin","TNM","Outcome",f"=B{r_tm}",f"=B{r_om}",f"=C{r_tm}",f"=C{r_om}",PCT)
hblk(13,"Price / SP","Break-even","Recommended",f"=B{r_be}",f"=B{r_rc}",f"=C{r_be}",f"=C{r_rc}",MONEY2)
for col in "NOP": ws.column_dimensions[col].hidden=True
def cbar(anchor,top,title,colors,pct=False):
    ch=BarChart(); ch.type="col"; ch.title=title; ch.height=6.6; ch.width=10.5; ch.style=10
    ch.add_data(Reference(ws,min_col=15,max_col=16,min_row=top,max_row=top+2),titles_from_data=True)
    ch.set_categories(Reference(ws,min_col=14,min_row=top+1,max_row=top+2))
    ch.legend.position="b"; ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True; ch.dataLabels.numFmt=('0.0%' if pct else '"$"#,##0')
    ch.series[0].graphicalProperties.solidFill=colors[0]; ch.series[1].graphicalProperties.solidFill=colors[1]; ch.y_axis.majorGridlines=None
    ws.add_chart(ch,anchor)
cbar("E5",5,"Client pays: TNM vs Outcome",[GREY,GREEN])
cbar("E20",9,"Opus margin %: TNM vs Outcome",[GREY,GREEN],pct=True)
cbar("E35",13,"Price per story point",[GREY,BLUE])
ws.freeze_panes="A3"

# ═══════════════════════════════ 5) DASHBOARD ═════════════════════════════════
ws=wb.create_sheet("Dashboard"); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor="C55A11"
for col,w in [("A",32),("B",18),("C",3),("D",24),("E",16)]: ws.column_dimensions[col].width=w
banner(ws,"E","Dashboard - TNM vs Outcome (Case 2, feasible)","Headline verdict for the 8 SP/person case.")
section(ws,4,"B","Headline numbers")
O=lambda ref: f"=Outcome!{ref}"
kpis=[("TNM monthly billing",O("B13"),MONEY,WHITE),("TNM yearly billing",O("B13")+"*12",MONEY,BG),
      ("TNM margin",O(f"C{r_tm}"),PCT,WHITE),("Outcome monthly bill",O(f"C{r_mb}"),MONEY,BG),
      ("Outcome margin",O(f"C{r_om}"),PCT,GREENL),("Opus extra profit / year",O(f"C{r_ep}"),MONEY,GREENL),
      ("Client saving / year",O(f"C{r_cy}"),MONEY,GREENL)]
for i,(lbl,frm,fmt,bf) in enumerate(kpis):
    rr=5+i
    setc(ws,f"A{rr}",lbl,f=font(11,NAVY,True),bg=bf,ha="left",border=True)
    big=bf==GREENL
    setc(ws,f"B{rr}",frm,f=font(14 if big else 12,(GREEN if big else INK),True),bg=bf,fmt=fmt,ha="center",border=True)
vr=5+len(kpis)+1
ws.merge_cells(f"A{vr}:B{vr+1}")
setc(ws,f"A{vr}",O(f"C{r_ww}"),f=font(16,WHITE,True),bg=GREEN,ha="center"); setc(ws,f"B{vr}",None,bg=GREEN)
for col in "AB":
    ws.cell(row=vr+1,column=1 if col=="A" else 2).fill=fill(GREEN)
ws.conditional_formatting.add(f"A{vr}",FormulaRule(formula=[f'ISNUMBER(SEARCH("WIN",A{vr}))'],fill=fill(GREEN),font=Font(color=WHITE,bold=True,size=16)))
ws.conditional_formatting.add(f"A{vr}",FormulaRule(formula=[f'NOT(ISNUMBER(SEARCH("WIN",A{vr})))'],fill=fill(AMBER),font=Font(color=WHITE,bold=True,size=16)))
# who benefits chart
setc(ws,"D4","Who benefits (annual $)",f=font(11,NAVY,True))
setc(ws,"D5","Client saving"); setc(ws,"E5",O(f"C{r_cy}"),fmt=MONEY)
setc(ws,"D6","Opus extra profit"); setc(ws,"E6",O(f"C{r_ep}"),fmt=MONEY)
ch=BarChart(); ch.type="col"; ch.title="Who benefits (annual $)"; ch.height=7.5; ch.width=11.5; ch.style=10
ch.add_data(Reference(ws,min_col=5,min_row=5,max_row=6)); ch.set_categories(Reference(ws,min_col=4,min_row=5,max_row=6))
ch.legend=None; ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True; ch.dataLabels.numFmt='"$"#,##0'
ch.series[0].graphicalProperties.solidFill=BLUE; ch.y_axis.majorGridlines=None
dp0=DataPoint(idx=0); dp0.graphicalProperties.solidFill=BLUE; ch.series[0].data_points.append(dp0)
dp1=DataPoint(idx=1); dp1.graphicalProperties.solidFill=GREEN; ch.series[0].data_points.append(dp1)
ws.add_chart(ch,"D8")
nr=vr+3
setc(ws,f"A{nr}","Both parties end up better off: the client pays less than TNM while Opus earns more - the efficiency Opus keeps under Outcome funds a price the client still finds cheaper.",f=font(11,NAVY,italic=True),ha="left",va="top",wrap=True)
ws.merge_cells(f"A{nr}:B{nr+5}")

# ═══════════════════════════════ VALIDATE & SAVE ══════════════════════════════
bad=[]
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for cl in row:
            v=cl.value
            if isinstance(v,str) and v.startswith("=") and any(ord(ch)>127 for ch in v):
                bad.append((sh.title,cl.coordinate))
if bad:
    print("NON-ASCII FORMULAS:",bad); raise SystemExit("abort")
print("Formula validation passed.")
OUT="Opus_Pulse_TNM_vs_Outcome_Model.xlsx"
try: wb.save(OUT); print("Saved",OUT)
except PermissionError:
    alt=OUT.replace(".xlsx","_UPDATED.xlsx"); wb.save(alt); print(OUT,"open - saved",alt)
